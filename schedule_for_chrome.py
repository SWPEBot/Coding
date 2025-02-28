import calendar
import datetime
import random
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side


WEEKDAYS = ["周天", "周一", "周二", "周三", "周四", "周五", "周六"]
BORDER_STYLE = Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))


def generate_schedule(year, month, staff_list):
    """
    Generate a schedule.

    Args:
        year: Year (int).
        month: Month (int).
        staff_list: List of staff (list of str).

    Returns:
        A dictionary where the key is the date (datetime.date object) and the value is the name of the staff member responsible for that date (str).
    """

    # Get_date
    start_date = datetime.date(year, month, 25)
    next_month = month + 1 if month < 12 else 1
    next_year = year + 1 if month == 12 else year
    end_date = datetime.date(next_year, next_month, 25)

    # List
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += datetime.timedelta(days=1)

    # Sort_list
    random.shuffle(staff_list)

    # Excel
    schedule = {}
    staff_index = 0
    for date in date_list:
        if date.weekday() < 5:
            schedule[date] = staff_list[staff_index]
            staff_index = (staff_index + 1) % len(staff_list)

    return schedule


def create_excel_schedule(schedule, year, month, filename="schedule.xlsx"):
    """
    Write the schedule to an Excel file, from the 25th of the current month to the 25th of the next month, arranged in a calendar format, only scheduling staff for weekdays.

    Args:
        schedule: Schedule (dict).
        year: Year (int).
        month: Month (int).
        filename: Excel file name (str).
    """

    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Schedule"

    # Calculate start and end dates
    start_date = datetime.date(year, month, 26)
    next_month = month + 1 if month < 12 else 1
    next_year = year + 1 if month == 12 else year
    end_date = datetime.date(next_year, next_month, 25)

    # Set header
    sheet["A1"] = f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')} Schedule"
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.merge_cells("A1:G1")  # Merge cells, header spans a week
    sheet["A1"].alignment = Alignment(horizontal="center")

    # Set weekday headers
    for col, day in enumerate(WEEKDAYS, start=1):
        sheet.cell(row=2, column=col).value = day
        sheet.cell(row=2, column=col).font = Font(bold=True)
        sheet.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    # Write dates and staff assignments
    row_num = 3
    current_date = start_date
    while current_date <= end_date:
        # Python's weekday() returns 0-6, where 0 is Monday and 6 is Sunday.
        # To match "Sun", "Mon", ..., adjust accordingly.
        col = (current_date.weekday() + 1) % 7 + 1  # Adjust column number, so 0 (Monday) becomes 2, 6 (Sunday) becomes 1

        sheet.cell(row=row_num, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row=row_num, column=col).border = BORDER_STYLE

        if current_date in schedule:
            staff = schedule[current_date]
            sheet.cell(row=row_num, column=col).value = f"{current_date.day}\n{staff}"  # Display date and staff
        else:
            sheet.cell(row=row_num, column=col).value = str(current_date.day)  # Only display date

        current_date += datetime.timedelta(days=1)

        # If it's Saturday, move to the next row
        if col == 7:
            row_num += 1

    # Auto-adjust column widths
    for col in range(1, 8):
        column_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 12

    # Save Excel file
    workbook.save(filename)
    print(f"Schedule saved to {filename}")


if __name__ == "__main__":
    # List of staff
    staff_list = ["高强", "王为", "李军辉", "靳宝成", "袁金鹏", "骆成功", "吴清松", "马英杰"]

    # Get current date
    now = datetime.datetime.now()
    year = now.year
    month = now.month

    # Generate schedule
    schedule = generate_schedule(year, month, staff_list)

    # Create Excel schedule
    output_directory = '/home/yang/Documents/python/'
    filename = f"{output_directory}schedule_{year}_{month}.xlsx"
    create_excel_schedule(schedule, year, month, filename)

